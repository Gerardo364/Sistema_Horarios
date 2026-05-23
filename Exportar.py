from fpdf import FPDF
from openpyxl import Workbook
class PDFHorario(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 8, 'LICEO ARMANDO REVERON - HORARIOS ACADEMICOS', 0, 1, 'C')
        self.ln(2)

def exportar_a_pdf(horario_maestro, nombre_archivo="horario_liceo.pdf"):
    pdf = PDFHorario('L', 'mm', 'A4')
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # 1. Agrupar el diccionario por Docente
    horarios_por_docente = {}
    for (dia, bloque, seccion), info in horario_maestro.items():
        docente = info['docente']
        if docente not in horarios_por_docente:
            horarios_por_docente[docente] = {}
        
        # Guardamos la sección y materia (Ej: "1A (Quím)")
        horarios_por_docente[docente][(dia, bloque)] = f"{seccion} ({info['materia'][:4].upper()})"

    dias = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]
    
    # Anchos basados en la foto (Columna de hora más estrecha, días más anchos)
    ancho_hora = 35  
    ancho_dia = 48   
    ancho_total = ancho_hora + (ancho_dia * 5)

    # 2. Dibujar la tabla idéntica a la foto para cada profesor
    for docente, agenda in horarios_por_docente.items():
        # --- ENCABEZADO DEL DOCENTE ---
        pdf.set_font("Arial", 'B', 10)
        pdf.set_fill_color(225, 225,225) # Gris claro como en la foto
        pdf.cell(ancho_total, 8, f"(PROF. {docente.upper()})", 1, 1, 'C', fill=True)

        # --- FILA DE LOS DÍAS ---
        pdf.set_font("Arial", 'B', 9)
        pdf.cell(ancho_hora, 8, "HORA", 1, 0, 'C', fill=True)
        for dia in dias:
            pdf.cell(ancho_dia, 8, dia.upper(), 1, 0, 'C', fill=True)
        pdf.ln()

        # --- ESPACIO PATRIÓTICO (7:50 am) ---
        pdf.set_font("Arial", 'B', 8)
        pdf.set_fill_color(240, 240, 240)
        pdf.cell(ancho_hora, 6, "7:50 am", 1, 0, 'C', fill=True)
        pdf.cell(ancho_dia * 5, 6, "ESPACIO PATRIOTICO", 1, 1, 'C', fill=True)

        # --- BLOQUE 1 (8:00 - 9:10) ---
        bloque_1 = "8:00-9:10"
        pdf.set_font("Arial", '', 9)
        pdf.cell(ancho_hora, 10, "8:00 am - 9:10 am", 1, 0, 'C')
        for dia in dias:
            texto = agenda.get((dia, bloque_1), "")
            pdf.cell(ancho_dia, 10, texto, 1, 0, 'C')
        pdf.ln()

        # --- RECESO 1 ---
        pdf.set_font("Arial", 'B', 8)
        pdf.cell(ancho_hora, 5, "9:10 am - 9:20 am", 1, 0, 'C', fill=True)
        pdf.cell(ancho_dia * 5, 5, "RECESO", 1, 1, 'C', fill=True)

        # --- BLOQUE 2 (9:20 - 10:30) ---
        bloque_2 = "9:20-10:30"
        pdf.set_font("Arial", '', 9)
        pdf.cell(ancho_hora, 10, "9:20 am - 10:30 am", 1, 0, 'C')
        for dia in dias:
            texto = agenda.get((dia, bloque_2), "")
            pdf.cell(ancho_dia, 10, texto, 1, 0, 'C')
        pdf.ln()

        # --- RECESO 2 ---
        pdf.set_font("Arial", 'B', 8)
        pdf.cell(ancho_hora, 5, "10:30 am - 10:35 am", 1, 0, 'C', fill=True)
        pdf.cell(ancho_dia * 5, 5, "RECESO", 1, 1, 'C', fill=True)

        # --- BLOQUE 3 (10:35 - 11:45) ---
        bloque_3 = "10:35-11:45"
        pdf.set_font("Arial", '', 9)
        pdf.cell(ancho_hora, 10, "10:35 am - 11:45 am", 1, 0, 'C')
        for dia in dias:
            texto = agenda.get((dia, bloque_3), "")
            pdf.cell(ancho_dia, 10, texto, 1, 0, 'C')
        pdf.ln()

        # --- RECESO 3 ---
        pdf.set_font("Arial", 'B', 8)
        pdf.cell(ancho_hora, 5, "11:45 am - 11:50 am", 1, 0, 'C', fill=True)
        pdf.cell(ancho_dia * 5, 5, "RECESO", 1, 1, 'C', fill=True)

        # --- BLOQUE 4 (11:50 - 13:00) ---
        bloque_4 = "11:50-13:00"
        pdf.set_font("Arial", '', 9)
        pdf.cell(ancho_hora, 10, "11:50 am - 1:00 pm", 1, 0, 'C')
        for dia in dias:
            texto = agenda.get((dia, bloque_4), "")
            pdf.cell(ancho_dia, 10, texto, 1, 0, 'C')
        pdf.ln()

        # Espacio final antes del siguiente profesor
        pdf.ln(8)

    pdf.output(nombre_archivo)
    print(f"PDF generado exitosamente en formato cartelera: {nombre_archivo}")


def exportar_a_excel(horario_maestro, nombre_archivo="horario_liceo.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    
    # Agrupar horario por docente (como en PDF)
    horarios_por_docente = {}
    for (dia, bloque, seccion), info in horario_maestro.items():
        docente = info['docente']
        if docente not in horarios_por_docente:
            horarios_por_docente[docente] = {}
        horarios_por_docente[docente][(dia, bloque)] = f"{seccion} ({info['materia'][:4].upper()})"
    
    dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    bloques = ["8:00-9:10", "9:20-10:30", "10:35-11:45", "11:50-13:00"]
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002046", end_color="002046", fill_type="solid")
    title_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for docente, agenda in horarios_por_docente.items():
        ws = wb.create_sheet(title=f"Prof. {docente[:31]}")  # Excel limita a 31 caracteres
        
        # Título del docente
        ws.merge_cells('A1:F1')
        ws['A1'] = f"PROF. {docente.upper()}"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_alignment
        
        # Encabezados de días
        headers = ["HORA", "LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Espacio patriótico (7:50 am)
        ws.merge_cells(f'A3:F3')
        ws['A3'] = "ESPACIO PATRIÓTICO (7:50 am)"
        ws['A3'].alignment = center_alignment
        ws['A3'].fill = PatternFill(start_color="f0f4fa", end_color="f0f4fa", fill_type="solid")
        
        row = 4
        for bloque in bloques:
            # Hora
            ws.cell(row=row, column=1, value=bloque).border = thin_border
            ws.cell(row=row, column=1).alignment = center_alignment
            
            # Días
            for i, dia in enumerate(dias, 2):
                valor = agenda.get((dia, bloque), "")
                cell = ws.cell(row=row, column=i, value=valor)
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Receso después de algunos bloques
            row += 1
            if bloque in ["8:00-9:10", "9:20-10:30", "10:35-11:45"]:
                ws.merge_cells(f'A{row}:F{row}')
                ws.cell(row=row, column=1, value="RECESO").alignment = center_alignment
                ws.cell(row=row, column=1).fill = PatternFill(start_color="efedf1", end_color="efedf1", fill_type="solid")
                row += 1
        
        # Ajustar anchos de columna
        for col in range(1, 7):
            ws.column_dimensions[chr(64+col)].width = 18
    
    # Eliminar hoja por defecto ("Sheet")
    wb.remove(wb["Sheet"])
    
    wb.save(nombre_archivo)
    print(f"Excel generado exitosamente: {nombre_archivo}")
    